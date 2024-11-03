
from common.debug import *
import openpyxl
from utils.logutil import logger
class ReadExcel():


    def read_xlsx_excel(self,excel_path, sheet_name):
        """
        :param url: 文件路径
        :param sheet_name: 文件中的sheet名称
        :return:  返回文件中的数据
        """
        # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
        workbook = openpyxl.load_workbook(excel_path)
        # 根据指定表名获取表格并得到对应的sheet对象
        sheet = workbook[sheet_name]
        print(sheet)
        # 定义列表存储表格数据
        data = []
        # 遍历表格的每一行
        for row in sheet.rows:
            # 定义表格存储每一行数据
            da = []
            # 从每一行中遍历每一个单元格
            for cell in row:
                # 将行数据存储到da列表
                da.append(cell.value)
            # 存储每一行数据
            data.append(da)
        # 返回数据
        return data




    def write_xlsx_excel_add(self,excel_path, sheet_name, two_dimensional_data):
        '''
        追加写入xlsx格式文件
        参数：
            excel_path:文件路径
            sheet_name:表名
            two_dimensional_data：将要写入表格的数据（二维列表）
        '''
        # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
        workbook = openpyxl.load_workbook(excel_path)
        # 根据指定表名获取表格并得到对应的sheet对象
        sheet = workbook[sheet_name]
        for tdd in two_dimensional_data:
            sheet.append(tdd)
        # 保存到指定位置
        workbook.save(excel_path)
        logger.info("数据追加写入成功。追加的数据是: {}".format(two_dimensional_data))


    def write_xlsx_excel(self,excel_path, sheet_name, two_dimensional_data):
        '''
        写入xlsx格式文件
        参数：
            excel_path:文件路径
            sheet_name:表名
            two_dimensional_data：将要写入表格的数据（二维列表）
        '''
        # 创建工作簿对象
        workbook = openpyxl.Workbook()
        # 创建工作表对象
        sheet = workbook.active
        # 设置该工作表的名字
        sheet.title = sheet_name
        # 遍历表格的每一行
        for i in range(0, len(two_dimensional_data)):
            # 遍历表格的每一列
            for j in range(0, len(two_dimensional_data[i])):
                # 写入数据（注意openpyxl的行和列是从1开始的，和我们平时的认知是一样的）
                sheet.cell(row=i + 1, column=j + 1, value=str(two_dimensional_data[i][j]))
        # 保存到指定位置
        workbook.save(excel_path)
        logger.info("数据写入成功。写入的数据是: {}".format(two_dimensional_data))



datas = [["万股",22,908,45112]]
biaotu = [["姓名","电话号码","电子邮件"]]
shuju = [
    [get_name(),get_phone(),get_email()]
]
read_excel = ReadExcel()
# data = read_excel.read_xlsx_excel("./datas/excel/ceshi.xlsx","测试")
# read_excel.write_xlsx_excel("./datas/excel/ceshisjiku.xlsx","xiaoy",biaotu)
for i in range(1000):
    shuju = [
        [get_name(), get_phone(), get_email()]
    ]
    read_excel.write_xlsx_excel_add("./datas/excel/ceshisjiku.xlsx","xiaoy",shuju)


# print(data)
#
# for i in data:
#     print(i)